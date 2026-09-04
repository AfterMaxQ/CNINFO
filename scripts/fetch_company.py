"""Normalize CNINFO node-company responses and export the nine-column XLSX."""

from __future__ import annotations

import argparse
import json
import re
from copy import copy
from pathlib import Path
from typing import Any, Iterable


HEADERS = ["主题", "信源主体", "分类1", "分类2", "分类3", "分类4", "公司", "信源URL", "备注"]
SOURCE_NAME = "CNINFO产业分析系统"
NOTE = "来自CNINFO产业链中心结构化数据"


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def response_data(payload: dict[str, Any]) -> dict[str, Any]:
    if payload.get("code") not in (None, 200, "200"):
        raise RuntimeError(f"CNINFO response failed: {payload.get('code')} {payload.get('msg')}")
    data = payload.get("data", payload)
    return data if isinstance(data, dict) else {}


def company_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    data = response_data(payload)
    rows = data.get("companys")
    if isinstance(rows, list):
        return rows
    rows = data.get("list")
    return rows if isinstance(rows, list) else []


def income_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    data = response_data(payload)
    rows = data.get("list")
    if isinstance(rows, dict):
        rows = rows.get("list")
    return rows if isinstance(rows, list) else []


def first_value(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def matching_product(row: dict[str, Any], industry_code: str) -> str:
    industries = row.get("industry") or []
    codes = row.get("industryCode") or []
    if isinstance(industries, list) and isinstance(codes, list) and industry_code in codes:
        index = codes.index(industry_code)
        return str(industries[index]) if index < len(industries) else ""
    return ""


def clean_name(value: str) -> str:
    return re.sub(r"\s+", "", value).replace("（", "(").replace("）", ")")


def add_company(store: dict[str, dict[str, Any]], name: str, code: str, listed: bool, source_type: str, product: str = "") -> None:
    name = name.strip()
    code = code.strip()
    if not name:
        return
    key = f"code:{code}" if code else f"name:{clean_name(name)}"
    if key not in store:
        store[key] = {
            "name": name,
            "code": code,
            "listed": listed,
            "source_types": [source_type],
            "products": [product] if product else [],
        }
        return
    existing = store[key]
    existing["listed"] = existing["listed"] or listed
    if source_type not in existing["source_types"]:
        existing["source_types"].append(source_type)
    if product and product not in existing["products"]:
        existing["products"].append(product)
    if len(name) > len(existing["name"]):
        existing["name"] = name


def normalize_companies(node_id: str, industry_code: str, node_name: str, income_payload: dict[str, Any], listed_payload: dict[str, Any], global_payloads: Iterable[dict[str, Any]]) -> dict[str, Any]:
    store: dict[str, dict[str, Any]] = {}
    for row in income_rows(income_payload):
        add_company(
            store,
            first_value(row, "company_name_one", "company_name_two"),
            first_value(row, "seccode_one", "seccode_two", "company_num_id_one", "company_num_id_two"),
            True,
            "annual_report_product",
            first_value(row, "product_name_one", "product_name_two"),
        )
    for row in company_rows(listed_payload):
        add_company(
            store,
            first_value(row, "fullname", "companyShortName"),
            first_value(row, "stockCode", "company_id"),
            True,
            "listed_search",
            matching_product(row, industry_code),
        )
    for global_payload in global_payloads:
        for row in company_rows(global_payload):
            add_company(
                store,
                first_value(row, "fullname", "companyShortName"),
                first_value(row, "stockCode", "company_id"),
                False,
                "non_listed_search",
                matching_product(row, industry_code),
            )

    companies = list(store.values())
    return {
        "node_id": node_id,
        "node": node_name,
        "industry_code": industry_code,
        "companies": companies,
        "counts": {
            "unique": len(companies),
            "listed": sum(1 for company in companies if company["listed"]),
            "non_listed": sum(1 for company in companies if not company["listed"]),
        },
        "source_types": ["annual_report_product", "listed_search", "non_listed_search"],
    }


def normalize_dom_capture(capture: dict[str, Any]) -> dict[str, Any]:
    """Normalize the structured company tables rendered by the signed-in page."""

    store: dict[str, dict[str, Any]] = {}
    for page in capture.get("responses", {}).get("listed_pages", []):
        for row in page.get("rows", []):
            add_company(store, first_value(row, "name"), first_value(row, "code"), True, "listed_page_table")
    for page in capture.get("responses", {}).get("non_listed_pages", []):
        for row in page.get("rows", []):
            add_company(store, first_value(row, "name"), first_value(row, "code"), False, "non_listed_page_table")
    companies = list(store.values())
    return {
        "node_id": capture.get("node_id", ""),
        "node": capture.get("node_name", ""),
        "industry_code": capture.get("industry_code", ""),
        "companies": companies,
        "counts": {
            "unique": len(companies),
            "listed": sum(1 for company in companies if company["listed"]),
            "non_listed": sum(1 for company in companies if not company["listed"]),
        },
        "pagination": capture.get("pagination", {}),
        "status": capture.get("status", ""),
        "capture_mode": capture.get("capture_mode", ""),
        "errors": capture.get("errors", []),
        "source_url": capture.get("source_url", ""),
    }


def iter_nodes(nodes: Iterable[dict[str, Any]], path: tuple[str, ...] = ()) -> Iterable[tuple[dict[str, Any], tuple[str, ...]]]:
    for node in nodes:
        current_path = path + (node.get("name", ""),)
        yield node, current_path
        yield from iter_nodes(node.get("children", []), current_path)


def build_xlsx(chain: dict[str, Any], companies: dict[str, Any], output: Path, source_url: str) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    company_names = "、".join(company["name"] for company in companies.get("companies", []))
    rows: list[list[str]] = []
    for node, path in iter_nodes(chain.get("tree", [])):
        node_companies = company_names if node.get("node_id") == companies.get("node_id") else ""
        direction = str(node.get("chain_updown") or "").strip()
        classification = [direction, path[0] if len(path) > 0 else "", path[1] if len(path) > 1 else "", " > ".join(path[2:]) if len(path) > 2 else ""]
        rows.append([
            str(chain.get("chain_name", "")),
            SOURCE_NAME,
            classification[0] if len(classification) > 0 else "",
            classification[1] if len(classification) > 1 else "",
            classification[2] if len(classification) > 2 else "",
            classification[3] if len(classification) > 3 else "",
            node_companies,
            source_url,
            NOTE if not rows else "",
        ])

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产业链"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        header_font = copy(cell.font)
        header_font.bold = True
        cell.font = header_font
    for row in sheet.iter_rows(min_row=2, min_col=8, max_col=8):
        cell = row[0]
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"
    widths = [18, 20, 24, 24, 24, 36, 70, 52, 36]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def build_all_xlsx(chain: dict[str, Any], companies_by_node: dict[str, dict[str, Any]], output: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产业链"
    sheet.append(HEADERS)
    for node, path in iter_nodes(chain.get("tree", [])):
        node_companies = companies_by_node.get(node.get("node_id", ""), {})
        company_names = "、".join(company["name"] for company in node_companies.get("companies", []))
        direction = str(node.get("chain_updown") or "").strip()
        classification = [direction, path[0] if len(path) > 0 else "", path[1] if len(path) > 1 else "", " > ".join(path[2:]) if len(path) > 2 else ""]
        source_url = node_companies.get("source_url") or chain.get("source_url", "")
        sheet.append([
            str(chain.get("chain_name", "")),
            SOURCE_NAME,
            classification[0],
            classification[1],
            classification[2],
            classification[3],
            company_names,
            source_url,
            NOTE if sheet.max_row == 2 else "",
        ])
    for cell in sheet[1]:
        header_font = copy(cell.font)
        header_font.bold = True
        cell.font = header_font
    for row in sheet.iter_rows(min_row=2, min_col=8, max_col=8):
        cell = row[0]
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"
    widths = [18, 20, 18, 24, 24, 36, 70, 52, 36]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def aggregate_dom_captures(chain: dict[str, Any], node_dir: Path, output: Path, xlsx_output: Path) -> dict[str, Any]:
    captures = [load_json(path) for path in sorted(node_dir.glob("*.json"))]
    by_node = {capture.get("node_id", ""): normalize_dom_capture(capture) for capture in captures if capture.get("node_id")}
    expected_nodes = list(iter_nodes(chain.get("tree", [])))
    missing = [node.get("node_id", "") for node, _ in expected_nodes if node.get("node_id", "") not in by_node]
    summary = {
        "nodes_expected": len(expected_nodes),
        "node_files": len(captures),
        "nodes_present": len(by_node),
        "nodes_missing": len(missing),
        "missing_node_ids": missing,
        "nodes_ok": sum(1 for item in by_node.values() if item.get("status") == "ok"),
        "nodes_no_industry_code": sum(1 for item in by_node.values() if item.get("status") == "no_industry_code"),
        "nodes_partial_or_error": sum(1 for item in by_node.values() if item.get("status") not in ("ok", "no_industry_code")),
        "listed_rows": sum(item["counts"]["listed"] for item in by_node.values()),
        "non_listed_rows": sum(item["counts"]["non_listed"] for item in by_node.values()),
    }
    result = {"chain_id": chain.get("chain_id", ""), "chain_name": chain.get("chain_name", ""), "summary": summary, "nodes": list(by_node.values())}
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    build_all_xlsx(chain, by_node, xlsx_output)
    return result


def main() -> None:
    project_dir = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--node-id", default="A02n019")
    parser.add_argument("--node-name", default="太阳能EVA胶膜")
    parser.add_argument("--industry-code", default="A02010201")
    parser.add_argument("--income-input", type=Path, default=project_dir / "data/raw/node_A02n019_companyIncome.json")
    parser.add_argument("--listed-input", type=Path, default=project_dir / "data/raw/node_A02n019_searchOtherListed.json")
    parser.add_argument("--global-input", type=Path, action="append")
    parser.add_argument("--output", type=Path, default=project_dir / "data/processed/company_A02n019.json")
    parser.add_argument("--chain-json", type=Path, default=project_dir / "data/processed/chain_lsx019.json")
    parser.add_argument("--xlsx-output", type=Path, default=project_dir / "export/result.xlsx")
    parser.add_argument("--source-url", default="https://pis.cninfo.com.cn/ics/index.html#/industryChain/A02n019/lsx019/A02n019/%E5%A4%AA%E9%98%B3%E8%83%BDEVA%E8%83%B6%E8%86%9C")
    parser.add_argument("--node-dir", type=Path)
    parser.add_argument("--all-output", type=Path, default=project_dir / "data/processed/newenergy_companies.json")
    parser.add_argument("--all-xlsx-output", type=Path, default=project_dir / "export/result.xlsx")
    args = parser.parse_args()

    if args.node_dir:
        aggregate = aggregate_dom_captures(load_json(args.chain_json), args.node_dir, args.all_output, args.all_xlsx_output)
        print(json.dumps({"chain": aggregate["chain_name"], "summary": aggregate["summary"], "json": str(args.all_output), "xlsx": str(args.all_xlsx_output)}, ensure_ascii=False))
        return

    global_inputs = args.global_input or sorted(project_dir.glob("data/raw/node_A02n019_searchglobalNew*.json"))
    if not global_inputs:
        raise FileNotFoundError("no searchglobalNew response files found")
    result = normalize_companies(
        args.node_id,
        args.industry_code,
        args.node_name,
        load_json(args.income_input),
        load_json(args.listed_input),
        (load_json(path) for path in global_inputs),
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    build_xlsx(load_json(args.chain_json), result, args.xlsx_output, args.source_url)
    print(json.dumps({"node": result["node"], "counts": result["counts"], "json": str(args.output), "xlsx": str(args.xlsx_output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
