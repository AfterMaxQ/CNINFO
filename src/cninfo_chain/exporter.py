from __future__ import annotations

import json
import os
import uuid
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from cninfo_chain.errors import CompanyCellTooLong, ExportLocked


HEADERS = (
    "主题",
    "信源主体",
    "分类1",
    "分类2",
    "分类3",
    "分类4",
    "公司",
    "信源URL",
    "备注",
)
SOURCE_NAME = "CNINFO产业分析系统"
SOURCE_REMARK = "来自CNINFO产业链中心结构化数据"


class XlsxExporter:
    def __init__(self, store: Any, target: Path) -> None:
        self.store = store
        self.target = Path(target)

    def export(self, run_id: str | None = None) -> Path:
        rows = self._business_rows(self.store.export_rows())
        self.target.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.target.with_name(
            f".{self.target.name}.{uuid.uuid4().hex}.tmp.xlsx"
        )
        try:
            self._write_workbook(temporary, rows)
            self._validate_workbook(temporary, len(rows))
            try:
                os.replace(temporary, self.target)
            except PermissionError as error:
                raise ExportLocked(
                    f"target XLSX is locked; close Excel and retry --export-now: {self.target}"
                ) from error
        finally:
            temporary.unlink(missing_ok=True)
        if run_id is not None:
            self.store.update_export_path(run_id, str(self.target))
        return self.target

    @staticmethod
    def _business_rows(source_rows: list[dict[str, Any]]) -> list[list[str]]:
        grouped: OrderedDict[int, list[dict[str, Any]]] = OrderedDict()
        for row in source_rows:
            grouped.setdefault(int(row["node_db_id"]), []).append(row)

        result: list[list[str]] = []
        seen_themes: set[str] = set()
        for node_rows in grouped.values():
            first = node_rows[0]
            raw_path = first["path_json"]
            path = json.loads(raw_path) if isinstance(raw_path, str) else list(raw_path)
            companies: list[str] = []
            seen_companies: set[str] = set()
            for row in node_rows:
                short_name = row.get("company_short_name")
                if short_name and short_name not in seen_companies:
                    seen_companies.add(short_name)
                    companies.append(str(short_name))
            company_cell = "、".join(companies)
            if len(company_cell) > 32_767:
                raise CompanyCellTooLong(
                    f"company cell exceeds Excel limit: {first['chain_name']} / {first['node_name']}"
                )
            chain_name = str(first["chain_name"])
            remark = SOURCE_REMARK if chain_name not in seen_themes else ""
            seen_themes.add(chain_name)
            result.append(
                [
                    chain_name,
                    SOURCE_NAME,
                    str(first["business_zone"]),
                    str(path[0]) if path else "",
                    str(path[1]) if len(path) > 1 else "",
                    " > ".join(str(item) for item in path[2:]),
                    company_cell,
                    str(first["source_url"]),
                    remark,
                ]
            )
        return result

    @staticmethod
    def _write_workbook(path: Path, rows: list[list[str]]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "产业链"
        sheet.append(HEADERS)
        for row in rows:
            sheet.append(row)
            url_cell = sheet.cell(sheet.max_row, 8)
            if url_cell.value:
                url_cell.hyperlink = str(url_cell.value)
                url_cell.style = "Hyperlink"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _validate_workbook(path: Path, expected_rows: int) -> None:
        workbook = load_workbook(path, read_only=False, data_only=False)
        try:
            sheet = workbook.active
            actual_headers = tuple(cell.value for cell in sheet[1])
            if actual_headers != HEADERS or sheet.max_column != len(HEADERS):
                raise ValueError("XLSX header contract validation failed")
            if sheet.max_row != expected_rows + 1:
                raise ValueError("XLSX row count validation failed")
            for row_number in range(2, sheet.max_row + 1):
                cell = sheet.cell(row_number, 8)
                if cell.value and (cell.hyperlink is None or cell.hyperlink.target != cell.value):
                    raise ValueError(f"XLSX URL hyperlink validation failed at row {row_number}")
        finally:
            workbook.close()
