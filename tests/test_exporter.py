from __future__ import annotations

import json

import pytest
from openpyxl import load_workbook

from cninfo_chain.errors import ExportLocked
from cninfo_chain.exporter import HEADERS, XlsxExporter


class FakeStore:
    def __init__(self, rows):
        self.rows = rows
        self.export_updates = []

    def export_rows(self):
        return list(self.rows)

    def update_export_path(self, run_id, path):
        self.export_updates.append((run_id, path))


def _rows():
    return [
        {
            "node_db_id": 1,
            "chain_name": "新能源",
            "chain_sort_no": 0,
            "node_name": "太阳能电池零部件",
            "business_zone": "上游",
            "node_sort_no": 0,
            "path_json": json.dumps(["太阳能电池零部件"], ensure_ascii=False),
            "source_url": "https://pis.cninfo.com.cn/parent",
            "company_sort_no": None,
            "company_name": None,
            "company_short_name": None,
            "listing_status": None,
        },
        {
            "node_db_id": 2,
            "chain_name": "新能源",
            "chain_sort_no": 0,
            "node_name": "太阳能EVA胶膜",
            "business_zone": "上游",
            "node_sort_no": 1,
            "path_json": ["太阳能电池零部件", "太阳能EVA胶膜"],
            "source_url": "https://pis.cninfo.com.cn/eva",
            "company_sort_no": 0,
            "company_name": "杭州福斯特应用材料股份有限公司",
            "company_short_name": "福斯特",
            "listing_status": 1,
        },
        {
            "node_db_id": 2,
            "chain_name": "新能源",
            "chain_sort_no": 0,
            "node_name": "太阳能EVA胶膜",
            "business_zone": "上游",
            "node_sort_no": 1,
            "path_json": ["太阳能电池零部件", "太阳能EVA胶膜"],
            "source_url": "https://pis.cninfo.com.cn/eva",
            "company_sort_no": 1,
            "company_name": "杭州福斯特应用材料股份有限公司",
            "company_short_name": "福斯特",
            "listing_status": 1,
        },
        {
            "node_db_id": 2,
            "chain_name": "新能源",
            "chain_sort_no": 0,
            "node_name": "太阳能EVA胶膜",
            "business_zone": "上游",
            "node_sort_no": 1,
            "path_json": ["太阳能电池零部件", "太阳能EVA胶膜"],
            "source_url": "https://pis.cninfo.com.cn/eva",
            "company_sort_no": 2,
            "company_name": "无锡市万力粘合材料股份有限公司",
            "company_short_name": None,
            "listing_status": 0,
        },
    ]


def test_exporter_writes_exact_nine_columns_one_row_per_node_and_short_names(tmp_path):
    target = tmp_path / "result.xlsx"
    store = FakeStore(_rows())
    result = XlsxExporter(store, target).export(run_id="run-1")

    workbook = load_workbook(result)
    sheet = workbook.active
    assert tuple(cell.value for cell in sheet[1]) == HEADERS
    assert sheet.max_column == 9
    assert sheet.max_row == 3
    assert [sheet.cell(2, column).value or "" for column in range(1, 10)] == [
        "新能源",
        "CNINFO产业分析系统",
        "上游",
        "太阳能电池零部件",
        "",
        "",
        "",
        "https://pis.cninfo.com.cn/parent",
        "来自CNINFO产业链中心结构化数据",
    ]
    assert sheet.cell(3, 7).value == "福斯特"
    assert "杭州福斯特应用材料股份有限公司" not in sheet.cell(3, 7).value
    assert sheet.cell(3, 8).hyperlink.target == "https://pis.cninfo.com.cn/eva"
    assert sheet.cell(3, 9).value is None
    assert store.export_updates == [("run-1", str(target))]


def test_locked_target_does_not_replace_existing_file(monkeypatch, tmp_path):
    target = tmp_path / "result.xlsx"
    target.write_bytes(b"existing")
    store = FakeStore(_rows())

    def locked(*_):
        raise PermissionError("locked by Excel")

    monkeypatch.setattr("cninfo_chain.exporter.os.replace", locked)
    with pytest.raises(ExportLocked):
        XlsxExporter(store, target).export(run_id="run-1")

    assert target.read_bytes() == b"existing"
    assert store.export_updates == []

